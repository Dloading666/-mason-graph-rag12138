"""Document text extraction and preview chunking."""

from __future__ import annotations

import json
import shutil
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

from fastapi import UploadFile

from backend.core.contracts import ChunkingConfig
from backend.search.chunker.mason_chunker import MasonChunker


SUPPORTED_TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".pdf", ".docx", ".doc", ".xlsx"}


class DocumentProcessor:
    """Extract UTF-8 text from the document formats used in enterprise knowledge bases."""

    def __init__(self) -> None:
        self.chunker = MasonChunker()

    async def extract_text(self, upload_file: UploadFile) -> str:
        suffix = Path(upload_file.filename or "").suffix.lower()
        if suffix not in SUPPORTED_TEXT_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {suffix or 'unknown'}")

        raw = await upload_file.read()
        if suffix in {".txt", ".md", ".csv"}:
            return raw.decode("utf-8", errors="ignore").strip()
        if suffix == ".json":
            payload = json.loads(raw.decode("utf-8", errors="ignore"))
            return json.dumps(payload, ensure_ascii=False, indent=2)
        if suffix in {".yaml", ".yml"}:
            try:
                import yaml
            except ImportError as exc:
                raise ValueError("YAML support requires PyYAML") from exc
            payload = yaml.safe_load(raw.decode("utf-8", errors="ignore"))
            return yaml.safe_dump(payload, allow_unicode=True, sort_keys=False).strip()
        if suffix == ".pdf":
            return self._extract_pdf(raw)
        if suffix == ".docx":
            return self._extract_docx(raw)
        if suffix == ".xlsx":
            return self._extract_xlsx(raw)
        if suffix == ".doc":
            return self._extract_doc(raw)
        return raw.decode("utf-8", errors="ignore").strip()

    def chunk_preview(self, text: str, config: ChunkingConfig | None = None) -> list[str]:
        if config is None:
            return self.chunker.chunk(text)
        return MasonChunker(config).chunk(text)

    def _extract_pdf(self, raw: bytes) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ValueError("PDF support requires pypdf") from exc
        reader = PdfReader(BytesIO(raw))
        return "\n".join((page.extract_text() or "").strip() for page in reader.pages).strip()

    def _extract_docx(self, raw: bytes) -> str:
        try:
            from docx import Document as DocxDocument
        except ImportError as exc:
            raise ValueError("DOCX support requires python-docx") from exc
        document = DocxDocument(BytesIO(raw))
        return "\n".join(paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip())

    def _extract_xlsx(self, raw: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX support requires openpyxl") from exc
        workbook = load_workbook(BytesIO(raw), data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append(" | ".join(values))
        return "\n".join(lines).strip()

    def _extract_doc(self, raw: bytes) -> str:
        antiword = shutil.which("antiword")
        if antiword is None:
            raise ValueError("Legacy .doc extraction requires antiword")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".doc") as temp_file:
            temp_file.write(raw)
            temp_path = Path(temp_file.name)

        try:
            result = subprocess.run(
                [antiword, str(temp_path)],
                capture_output=True,
                check=True,
                text=True,
                encoding="utf-8",
                errors="ignore",
            )
            return result.stdout.strip()
        finally:
            temp_path.unlink(missing_ok=True)
